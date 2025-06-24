from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Order, OrderStatus, OrderComment
from .forms import OrderForm, OrderEditForm, OrderCommentForm

@login_required
def order_list(request):
    """Список всех заявок"""
    orders = Order.objects.select_related('customer', 'manager', 'status').all()
    
    # Поиск
    search = request.GET.get('search')
    if search:
        orders = orders.filter(
            Q(number__icontains=search) |
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(customer__name__icontains=search)
        )
    
    # Фильтрация по статусу
    status_filter = request.GET.get('status')
    if status_filter:
        orders = orders.filter(status__name=status_filter)
    
    # Фильтрация по приоритету
    priority_filter = request.GET.get('priority')
    if priority_filter:
        orders = orders.filter(priority=priority_filter)
    
    # Фильтрация по менеджеру
    manager_filter = request.GET.get('manager')
    if manager_filter:
        orders = orders.filter(manager__username=manager_filter)
    
    # Статистика
    total_orders = orders.count()
    new_orders = orders.filter(status__name=OrderStatus.NEW).count()
    in_progress_orders = orders.filter(status__name=OrderStatus.IN_PROGRESS).count()
    completed_orders = orders.filter(status__name=OrderStatus.COMPLETED).count()
    overdue_orders = sum(1 for order in orders if order.is_overdue)
    
    # Пагинация
    paginator = Paginator(orders, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'manager_filter': manager_filter,
        'total_orders': total_orders,
        'new_orders': new_orders,
        'in_progress_orders': in_progress_orders,
        'completed_orders': completed_orders,
        'overdue_orders': overdue_orders,
        'statuses': OrderStatus.objects.all(),
        'priorities': [
            ('low', 'Низкий'),
            ('medium', 'Средний'),
            ('high', 'Высокий'),
            ('urgent', 'Срочный'),
        ],
    }
    return render(request, 'orders/order_list.html', context)

@login_required
def order_create(request):
    """Создание новой заявки"""
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.manager = request.user
            # Устанавливаем статус "Новая" по умолчанию
            new_status = OrderStatus.objects.get(name=OrderStatus.NEW)
            order.status = new_status
            order.save()
            messages.success(request, 'Заявка успешно создана!')
            return redirect('orders:order_detail', pk=order.pk)
    else:
        form = OrderForm()
    
    return render(request, 'orders/order_create.html', {'form': form})

@login_required
def order_detail(request, pk):
    """Детальная информация о заявке"""
    order = get_object_or_404(Order, pk=pk)
    comments = order.comments.all()
    
    if request.method == 'POST':
        comment_form = OrderCommentForm(request.POST)
        if comment_form.is_valid():
            comment = comment_form.save(commit=False)
            comment.order = order
            comment.author = request.user
            comment.save()
            messages.success(request, 'Комментарий добавлен!')
            return redirect('orders:order_detail', pk=pk)
    else:
        comment_form = OrderCommentForm()
    
    context = {
        'order': order,
        'comments': comments,
        'comment_form': comment_form,
    }
    return render(request, 'orders/order_detail.html', context)

@login_required
def order_edit(request, pk):
    """Редактирование заявки"""
    order = get_object_or_404(Order, pk=pk)
    
    if request.method == 'POST':
        form = OrderEditForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            messages.success(request, 'Заявка успешно обновлена!')
            return redirect('orders:order_detail', pk=pk)
    else:
        form = OrderEditForm(instance=order)
    
    return render(request, 'orders/order_edit.html', {'form': form, 'order': order})

@login_required
def order_delete(request, pk):
    """Удаление заявки"""
    order = get_object_or_404(Order, pk=pk)
    
    if request.method == 'POST':
        order.delete()
        messages.success(request, 'Заявка успешно удалена!')
        return redirect('orders:order_list')
    
    return render(request, 'orders/order_delete.html', {'order': order})

@login_required
def add_comment(request, pk):
    """Добавление комментария к заявке"""
    order = get_object_or_404(Order, pk=pk)
    
    if request.method == 'POST':
        form = OrderCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.order = order
            comment.author = request.user
            comment.save()
            messages.success(request, 'Комментарий добавлен!')
    
    return redirect('orders:order_detail', pk=pk)

@login_required
def generate_report(request):
    """Генерация отчета по заявкам"""
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            
            orders = Order.objects.filter(
                created_at__range=[start_date, end_date]
            ).select_related('customer', 'manager', 'status')
            
            # Создаем Excel файл
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет по заявкам"
            
            # Заголовки
            headers = ['Номер', 'Клиент', 'Дата создания', 'Статус', 'Ответственный менеджер', 'Заголовок', 'Приоритет']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Данные
            for row, order in enumerate(orders, 2):
                ws.cell(row=row, column=1, value=order.number)
                ws.cell(row=row, column=2, value=order.customer.name)
                ws.cell(row=row, column=3, value=order.created_at.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row, column=4, value=order.status.get_name_display())
                ws.cell(row=row, column=5, value=order.manager.username if order.manager else 'Не назначен')
                ws.cell(row=row, column=6, value=order.title)
                ws.cell(row=row, column=7, value=order.get_priority_display())
            
            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=orders_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
            
            wb.save(response)
            return response
    
    return render(request, 'orders/generate_report.html')
