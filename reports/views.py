from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from accounts.utils import role_required
from orders.models import Order
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta

# Create your views here.

@login_required
@role_required('admin', 'manager')
def report_orders(request):
    report_data = None
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if start_date and end_date:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            orders = Order.objects.filter(created_at__range=[start_dt, end_dt]).select_related('customer', 'manager', 'status')
            # Генерация Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет по заявкам"
            headers = ['Номер', 'Клиент', 'Дата создания', 'Статус', 'Менеджер']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for row, order in enumerate(orders, 2):
                ws.cell(row=row, column=1, value=order.number)
                ws.cell(row=row, column=2, value=order.customer.name)
                ws.cell(row=row, column=3, value=order.created_at.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row, column=4, value=order.status.get_name_display())
                ws.cell(row=row, column=5, value=order.manager.username if order.manager else '—')
            for col in range(1, 6):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f"orders_report_{start_date}_to_{end_date}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
    return render(request, 'reports/report_orders.html', {})
